from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QFileDialog
from design import report_prog
import sys
import csv
import os
import subprocess
import time

# import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

#---------------------------------------------------------------------------
from PyQt5.QtWidgets import QDialog, QApplication, QPushButton, QVBoxLayout
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import matplotlib.pyplot as plt
#---------------------------------------------------------------------------

class Window(QtWidgets.QMainWindow):
    def __init__(self):
        super(Window, self).__init__()      # Подключение интерфейса программы
        self.ui = report_prog.Ui_MainWindow()
        self.ui.setupUi(self)

        self.ui.pushButton.clicked.connect(self.btnOpen)
        self.ui.pushButton_2.clicked.connect(self.btnSave)
        self.ui.comboBox.activated.connect(self.idSorting)
        self.ui.comboBox_2.activated.connect(self.dataParsing)

    # ---------------------------------------------------------------------------
        self.ui.figure = plt.figure()       # Создаем фигуру графика
        self.ui.canvas = FigureCanvas(self.ui.figure)   # Поле для графика
        #self.ui.toolbar = NavigationToolbar(self.ui.canvas, self)   # Панель управления графика
        layout = QVBoxLayout()  # создаем контейнер и добавляем в него поле и тулбар
        #layout.addWidget(self.ui.toolbar)
        layout.addWidget(self.ui.canvas)

        self.ui.widget.setLayout(layout)    # добавление контейнера в интерфейс
        self.directory = ""
        self.data = [[],[],[],[]]   # [№], [температура], [время], [стадия]


    def btnOpen(self):  # Кнопка открыть
        #fname = QFileDialog.getOpenFileName(self, 'Открыть файл', '/home', '*.csv;;')[0]
        self.directory = QFileDialog.getExistingDirectory(None, 'Select a folder:', 'C:\\', QFileDialog.ShowDirsOnly)
        self.fileSorting()

    def btnSave(self):  # Кнопка сохранить
        fname = QFileDialog.getSaveFileName(self, 'Открыть файл', '/report', '*.xlsx')[0]
        # print(fname)
        self.exportToExcel(fname)

    def fileSorting(self):   # Берем в директории + /ID файлы и отправляем в выпадающий список

        try:
            directoryId = self.directory + "/ID"
            files = os.listdir(directoryId)
            for file in files:
                if '.dtl' in file:
                    if file.rstrip('dtl')+'xlsx' in files:
                        pass
                    else:
                        subprocess.Popen(('start', self.directory + "/ID/" + file), shell=True)
                        subprocess.Popen(('start', self.directory + "/Temp/" + file), shell=True)
                        subprocess.Popen(('start', self.directory + "/Stage/" + file), shell=True)
                else:
                    pass
            time.sleep(1)
            files = os.listdir(directoryId)
            # for file in files:
            #     if '.xlsx' in file:
            #         if file.rstrip('xlsx')+'csv' in files:
            #             pass
            #         else:
            #             self.csv_from_excel(self.directory +"/ID/"+file)
            #     else:
            #         pass
            hystoryFiles = []
            for file in files:
                if ('.csv' in file)or('.xlsx' in file):
                    hystoryFiles.append(file)
            self.ui.comboBox.clear()
            self.ui.comboBox.addItems(hystoryFiles)
            self.idSorting()
        except Exception:
            pass

    def csv_from_excel(self, file):

        try:
            xlsx = openpyxl.load_workbook(file)
            sheet = xlsx.active
            data = sheet.rows
            fileCSV = file.rstrip('xlsx')+'csv'
            csv = open(fileCSV, "w+")

            dataForCsv = []
            countRow = 0

            for row in data:
                dataForCsv.append([])
                for cell in row:
                    dataForCsv[countRow].append(str(cell.value))
                countRow = countRow+1

            for x in dataForCsv:
                x.pop(2)
                s = ",".join(x)
                csv.writelines(s)
                csv.write('\n')

            ## close the csv file
            csv.close()
        except Exception:
            pass



    def idSorting(self):
        try:
            directoryId = self.directory + "/ID/" + self.ui.comboBox.currentText()
            hystoryList = []
            if '.csv' in directoryId:
                fileId = open(directoryId, encoding="utf-8")

                readerId = csv.reader(fileId, delimiter=',')
                count = 0

                for x in readerId:
                    if count == 0:
                        pass
                    else:
                        if not x[2] in hystoryList:
                            hystoryList.append(x[2])
                    count += 1
                fileId.close()

            else:
                xlsx = openpyxl.load_workbook(directoryId)
                sheet = xlsx.active
                data = sheet.rows

                countRow = 0

                for row in data:
                    if countRow == 0:
                        print('1')
                        pass
                    else:
                        if not str(row[3].value) in hystoryList:
                            hystoryList.append(str(row[3].value))
                    countRow += 1

            self.ui.comboBox_2.clear()
            self.ui.comboBox_2.addItems(hystoryList)
        except Exception:
            pass


        # print("x")

        # self.dataParsing()

    def dataParsing(self):
        try:
            dirId = self.directory + "/ID/" + self.ui.comboBox.currentText()
            dirTemp = self.directory + "/Temp/" + self.ui.comboBox.currentText()
            dirStage = self.directory + "/Stage/" + self.ui.comboBox.currentText()

            self.data = [[], [], [], []]

            if '.csv' in self.ui.comboBox.currentText():
                fileId = open(dirId, encoding = "utf-8")
                fileTemp = open(dirTemp, encoding="utf-8")
                fileStage = open(dirStage, encoding="utf-8")

                readerId = csv.reader(fileId)
                readerTemp = csv.reader(fileTemp)
                readerStage = csv.reader(fileStage)

                listId = []
                listTemp = []
                listStage = []
                for x in readerId:
                    listId.append(x)
                for x in readerTemp:
                    listTemp.append(x)
                for x in readerStage:
                    listStage.append(x)

                count = 0

                for i in range(len(listId)):
                    if listId[i][2]==self.ui.comboBox_2.currentText():
                        self.data[0].append(count)
                        count+=1
                        self.data[1].append(listTemp[i][2])
                        self.data[2].append(listId[i][1])
                        self.data[3].append(listStage[i][2])

            else:
                xlsxId = openpyxl.load_workbook(dirId)
                xlsxTemp = openpyxl.load_workbook(dirTemp)
                xlsxStage = openpyxl.load_workbook(dirStage)

                sheetId = xlsxId.active
                sheetTemp = xlsxTemp.active
                sheetStage = xlsxStage.active

                dataId = sheetId.rows
                dataTemp = sheetTemp.rows
                dataStage = sheetStage.rows

                listIdXl = []
                listTempXl = []
                listStageXl = []
                for x in dataId:
                    listIdXl.append(x)
                for x in dataTemp:
                    listTempXl.append(x)
                for x in dataStage:
                    listStageXl.append(x)

                countRow = 0

                for i in range(len(listIdXl)):
                    if str(listIdXl[i][3].value) == self.ui.comboBox_2.currentText():
                        self.data[0].append(countRow)
                        countRow += 1
                        self.data[1].append(str(listTempXl[i][3].value))
                        self.data[2].append(str(listIdXl[i][1].value))
                        self.data[3].append(str(listStageXl[i][3].value))

            for x in self.data:
                print(x)
        except Exception:
            pass

    def exportToExcel(self, dir):
        try:
            wb = Workbook()
            wb.create_sheet(title="Отчет", index=0)
            sheet = wb["Отчет"]

            sheet['A1'] = "Отчет"
            sheet['B1'] = "ID: "+self.ui.comboBox_2.currentText()

            for x in range(len(self.data[0])):
                cell = sheet.cell(row = x+24, column=1)
                cell.value = self.data[0][x]
                cell = sheet.cell(row=x + 24, column=2)
                cell.value = float(self.data[1][x])
                cell = sheet.cell(row=x + 24, column=3)
                cell.value = self.data[2][x]
                cell = sheet.cell(row=x + 24, column=4)
                cell.value = self.data[3][x]

            chart = LineChart()
            chart.title = "Отчет"
            dataChart = Reference(sheet, min_col=2, min_row=24, max_col=2, max_row=len(self.data[0])+23)

            chart.add_data(dataChart, titles_from_data=True)
            sheet.add_chart(chart, 'A5')
            wb.save(dir)
        except Exception:
            pass












    # def plot(self): # Строим график
    #     # random data
    #     data = [random.random() for i in range(10)]
    #
    #     # clearing old figure
    #     self.ui.figure.clear()
    #
    #     # create an axis
    #     ax = self.ui.figure.add_subplot(111)
    #
    #     # plot data
    #     ax.plot(data, '*-')
    #
    #     # refresh canvas
    #     self.ui.canvas.draw()



app = QtWidgets.QApplication([])
application = Window()
application.show()

sys.exit(app.exec())
